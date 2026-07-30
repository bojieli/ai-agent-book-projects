"""Real data, webhook, and browser execution tools for Experiment 4-2."""

from __future__ import annotations

import hashlib
import json
import os
import shutil
import subprocess
import time
from pathlib import Path
from typing import Any

import httpx
from config import Config


def _sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _safe_output(path: str) -> Path:
    candidate = Path(path)
    if not candidate.is_absolute():
        candidate = Path(Config.WORKSPACE_DIR) / candidate
    candidate = candidate.resolve()
    candidate.relative_to(Path(Config.WORKSPACE_DIR).resolve())
    candidate.parent.mkdir(parents=True, exist_ok=True)
    return candidate


class ExtendedTools:
    async def excel_create_with_formula_and_screenshot(
        self, output_path: str, rows: list[dict[str, Any]]
    ) -> dict[str, Any]:
        """Create a real XLSX, apply formulas, and render a screenshot via LibreOffice."""
        from openpyxl import Workbook

        target = _safe_output(output_path)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Invoice"
        sheet.append(["Item", "Quantity", "Unit price", "Total"])
        for index, row in enumerate(rows, 2):
            sheet.append([row["item"], float(row["quantity"]), float(row["unit_price"]),
                          f"=B{index}*C{index}"])
        total_row = len(rows) + 2
        sheet.cell(total_row, 3, "Grand total")
        sheet.cell(total_row, 4, f"=SUM(D2:D{total_row - 1})")
        sheet.freeze_panes = "A2"
        sheet.column_dimensions["A"].width = 28
        for column in ("B", "C", "D"):
            sheet.column_dimensions[column].width = 16
        workbook.save(target)

        soffice = shutil.which("soffice") or shutil.which("libreoffice")
        if not soffice:
            return {"success": False, "error": "LibreOffice is required for formula rendering"}
        started = time.perf_counter()
        process = subprocess.run(
            [soffice, "--headless", "--convert-to", "pdf", "--outdir",
             str(target.parent), str(target)],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=120,
        )
        pdf = target.with_suffix(".pdf")
        if process.returncode != 0 or not pdf.is_file():
            return {"success": False, "error": process.stderr or process.stdout,
                    "returncode": process.returncode}
        import fitz

        document = fitz.open(pdf)
        screenshot = target.with_suffix(".png")
        document[0].get_pixmap(matrix=fitz.Matrix(1.5, 1.5), alpha=False).save(screenshot)
        document.close()
        return {
            "success": True,
            "xlsx": {"path": str(target), "bytes": target.stat().st_size, "sha256": _sha(target)},
            "pdf": {"path": str(pdf), "bytes": pdf.stat().st_size, "sha256": _sha(pdf)},
            "screenshot": {"path": str(screenshot), "bytes": screenshot.stat().st_size,
                           "sha256": _sha(screenshot)},
            "formula_cells": [f"D{index}" for index in range(2, total_row + 1)],
            "rows": len(rows),
            "renderer": "LibreOffice headless + PyMuPDF",
            "latency_seconds": round(time.perf_counter() - started, 3),
        }

    async def webhook_post(self, url: str, payload: dict[str, Any]) -> dict[str, Any]:
        """POST JSON to a real HTTPS webhook and retain response evidence."""
        if not url.startswith("https://"):
            return {"success": False, "error": "Only HTTPS webhook URLs are allowed"}
        started = time.perf_counter()
        async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
            response = await client.post(url, json=payload)
        try:
            body = response.json()
        except ValueError:
            body = {"text": response.text[:2000]}
        return {
            "success": response.is_success,
            "status": response.status_code,
            "url": str(response.url),
            "response": body,
            "response_sha256": hashlib.sha256(response.content).hexdigest(),
            "response_bytes": len(response.content),
            "latency_seconds": round(time.perf_counter() - started, 3),
        }

    async def browser_navigate(self, url: str, screenshot_path: str) -> dict[str, Any]:
        """Navigate with real headless Chromium, extract content, and retain pixels."""
        if not url.startswith("https://"):
            return {"success": False, "error": "Only HTTPS URLs are allowed"}
        target = _safe_output(screenshot_path)
        started = time.perf_counter()
        from playwright.async_api import async_playwright

        async with async_playwright() as playwright:
            browser = await playwright.chromium.launch(headless=True)
            page = await browser.new_page(viewport={"width": 1280, "height": 720})
            response = await page.goto(url, wait_until="networkidle", timeout=60000)
            title = await page.title()
            text = (await page.locator("body").inner_text())[:4000]
            await page.screenshot(path=str(target), full_page=True)
            await browser.close()
        return {
            "success": bool(response and response.ok and target.is_file()),
            "url": url,
            "status": response.status if response else None,
            "title": title,
            "body_text": text,
            "screenshot": {"path": str(target), "bytes": target.stat().st_size,
                           "sha256": _sha(target)},
            "browser": "Chromium via Playwright",
            "latency_seconds": round(time.perf_counter() - started, 3),
        }

    async def environment_capabilities(self) -> dict[str, Any]:
        """Report, without simulation, whether desktop/mobile backends are actually usable."""
        docker_image = subprocess.run(
            ["docker", "image", "inspect",
             "ghcr.io/anthropics/anthropic-quickstarts:computer-use-demo-latest"],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        ).returncode == 0 if shutil.which("docker") else False
        adb = shutil.which("adb")
        devices = ""
        if adb:
            devices = subprocess.run([adb, "devices"], stdout=subprocess.PIPE,
                                     stderr=subprocess.STDOUT, text=True).stdout
        active_devices = [line for line in devices.splitlines()[1:] if line.strip().endswith("device")]
        return {
            "success": True,
            "computer_use_container_image_present": docker_image,
            "computer_use_active_session": False,
            "android_world_adb_present": bool(adb),
            "android_active_devices": active_devices,
            "note": "Availability probe only; absent active sessions cannot satisfy execution gates.",
        }
